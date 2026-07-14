"""Kapsamlı birim + örnek Excel testleri."""

from __future__ import annotations

import tempfile
import unittest
from datetime import datetime, timedelta
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook

from analyzer import (
    analyze_workbook,
    build_report_workbook,
    format_seconds,
    format_text_summary,
    is_valid_personnel,
    normalize_personnel_name,
    normalize_phone,
    parse_duration_to_seconds,
    tr_lower,
    tr_upper,
)
from bot import is_group_chat, safe_filename

SAMPLE = Path(__file__).with_name("048c507c-32fc-4c7e-b7b6-17fc5d8915ad.xlsx")


def _make_xlsx(rows: list[tuple]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.append(
        [
            "TELEFON",
            "ARAMA TARİHİ",
            "ARAMA SAATİ",
            "CallID",
            "KONUŞMA SÜRESİ",
            "ÇALDIRMA SÜRESİ",
            "DAHİLİ ADI",
        ]
    )
    for r in rows:
        ws.append(list(r))
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


class TestPersonnelFilter(unittest.TestCase):
    def test_valid_names(self):
        self.assertTrue(is_valid_personnel("seda  -O"))
        self.assertTrue(is_valid_personnel("Ahmet-O"))
        self.assertTrue(is_valid_personnel("celal  -O"))
        self.assertTrue(is_valid_personnel("Seda-o"))

    def test_invalid_names(self):
        self.assertFalse(is_valid_personnel("-K"))
        self.assertFalse(is_valid_personnel("K"))
        self.assertFalse(is_valid_personnel("-"))
        self.assertFalse(is_valid_personnel("  -O"))
        self.assertFalse(is_valid_personnel(None))
        self.assertFalse(is_valid_personnel(""))
        self.assertFalse(is_valid_personnel("   "))

    def test_normalize_name(self):
        self.assertEqual(normalize_personnel_name("seda  -O"), "Seda")
        self.assertEqual(normalize_personnel_name("CELAL  -O"), "Celal")

    def test_turkish_english_chars(self):
        # Aynı kişi farklı yazımlar → tek isim
        self.assertEqual(normalize_personnel_name("İREM  -O"), "İrem")
        self.assertEqual(normalize_personnel_name("irem -O"), "İrem")
        self.assertEqual(normalize_personnel_name("İrem-O"), "İrem")
        self.assertEqual(normalize_personnel_name("ŞULE  -O"), "Şule")
        self.assertEqual(normalize_personnel_name("şule -o"), "Şule")
        self.assertEqual(normalize_personnel_name("GÖKÇE -O"), "Gökçe")
        self.assertEqual(normalize_personnel_name("gökçe -O"), "Gökçe")
        self.assertEqual(normalize_personnel_name("IŞIL -O"), "Işıl")
        # İngilizce
        self.assertEqual(normalize_personnel_name("JOHN -O"), "John")
        self.assertEqual(normalize_personnel_name("john -O"), "John")
        # tr_lower / tr_upper (Türkçe: I↔ı, İ↔i)
        self.assertEqual(tr_lower("IĞDIR"), "ığdır")
        self.assertEqual(tr_lower("İĞDIR"), "iğdır")
        self.assertEqual(tr_upper("ığdır"), "IĞDIR")
        self.assertEqual(tr_upper("iğdır"), "İĞDIR")
        self.assertTrue(is_valid_personnel("şule  -O"))
        self.assertTrue(is_valid_personnel("İREM -O"))

    def test_turkish_merge_in_workbook(self):
        """Farklı büyük/küçük yazımlar aynı personelde birleşir."""
        buf = _make_xlsx(
            [
                ("905777777777", "13.07.2026", "10:00:00", "x", "00:00:01", "00:00:01", "İREM  -O"),
                ("905777777777", "13.07.2026", "10:01:00", "x", "00:00:01", "00:00:01", "irem  -O"),
                ("905777777777", "13.07.2026", "10:02:00", "x", "00:00:01", "00:00:01", "İrem -O"),
            ]
        )
        reports = analyze_workbook(buf, min_repeat=2)
        self.assertEqual(len(reports), 1)
        self.assertEqual(reports[0].personel, "İrem")
        self.assertEqual(reports[0].tekrarli_numaralar[0].arama_sayisi, 3)


class TestPhoneAndDuration(unittest.TestCase):
    def test_phone_ok(self):
        self.assertEqual(normalize_phone("905307832114"), "905307832114")
        self.assertEqual(normalize_phone(905307832114), "905307832114")
        self.assertEqual(normalize_phone(905307832114.0), "905307832114")

    def test_phone_reject(self):
        self.assertEqual(normalize_phone("v"), "")
        self.assertEqual(normalize_phone("false"), "")
        self.assertEqual(normalize_phone("123"), "")
        self.assertEqual(normalize_phone(None), "")
        self.assertEqual(normalize_phone(True), "")

    def test_duration(self):
        self.assertEqual(parse_duration_to_seconds("00:00:56"), 56)
        self.assertEqual(parse_duration_to_seconds("00:01:30"), 90)
        self.assertEqual(parse_duration_to_seconds("false"), 0)
        self.assertEqual(parse_duration_to_seconds(timedelta(seconds=12)), 12)
        self.assertEqual(parse_duration_to_seconds(datetime(1899, 12, 30, 0, 0, 8)), 8)
        self.assertEqual(format_seconds(3661), "01:01:01")


class TestAnalysisLogic(unittest.TestCase):
    def test_personnel_isolation(self):
        """Aynı numara Ahmet ve Mehmet tarafından aransa ayrı sayılır."""
        buf = _make_xlsx(
            [
                ("905111111111", "13.07.2026", "10:00:00", "x", "00:00:10", "00:00:02", "ahmet  -O"),
                ("905111111111", "13.07.2026", "10:05:00", "x", "00:00:05", "00:00:01", "ahmet  -O"),
                ("905111111111", "13.07.2026", "11:00:00", "x", "00:00:03", "00:00:01", "mehmet  -O"),
                ("905111111111", "13.07.2026", "11:10:00", "x", "00:00:04", "00:00:01", "mehmet  -O"),
                ("905111111111", "13.07.2026", "12:00:00", "x", "00:00:01", "00:00:01", "ayse  -O"),  # 1x
            ]
        )
        reports = {r.personel: r for r in analyze_workbook(buf, min_repeat=2)}
        self.assertIn("Ahmet", reports)
        self.assertIn("Mehmet", reports)
        self.assertIn("Ayse", reports)
        self.assertEqual(len(reports["Ahmet"].tekrarli_numaralar), 1)
        self.assertEqual(reports["Ahmet"].tekrarli_numaralar[0].arama_sayisi, 2)
        self.assertEqual(reports["Mehmet"].tekrarli_numaralar[0].arama_sayisi, 2)
        # Ayşe 1 kez aradı → tekrarda yok
        self.assertEqual(len(reports["Ayse"].tekrarli_numaralar), 0)
        self.assertEqual(reports["Ayse"].toplam_arama, 1)

    def test_exclude_k_rows(self):
        buf = _make_xlsx(
            [
                ("905222222222", "13.07.2026", "10:00:00", "x", "00:00:00", "00:00:01", "-K"),
                ("905222222222", "13.07.2026", "10:01:00", "x", "00:00:00", "00:00:01", "-K"),
                ("905222222222", "13.07.2026", "10:02:00", "x", "00:00:00", "00:00:01", "seda  -O"),
                ("905222222222", "13.07.2026", "10:03:00", "x", "00:00:00", "00:00:01", "seda  -O"),
            ]
        )
        reports = analyze_workbook(buf, min_repeat=2)
        names = {r.personel for r in reports}
        self.assertNotIn("K", names)
        self.assertEqual(len(reports), 1)
        self.assertEqual(reports[0].personel, "Seda")
        self.assertEqual(reports[0].tekrarli_numaralar[0].arama_sayisi, 2)

    def test_duration_totals(self):
        buf = _make_xlsx(
            [
                ("905333333333", "13.07.2026", "10:00:00", "x", "00:00:10", "00:00:05", "ali  -O"),
                ("905333333333", "13.07.2026", "10:01:00", "x", "00:00:20", "00:00:07", "ali  -O"),
            ]
        )
        rep = analyze_workbook(buf, min_repeat=2)[0]
        n = rep.tekrarli_numaralar[0]
        self.assertEqual(n.toplam_konusma_saniye, 30)
        self.assertEqual(n.toplam_caldirma_saniye, 12)

    def test_min_repeat_threshold(self):
        buf = _make_xlsx(
            [
                ("905444444444", "13.07.2026", "10:00:00", "x", "00:00:01", "00:00:01", "can  -O"),
                ("905444444444", "13.07.2026", "10:01:00", "x", "00:00:01", "00:00:01", "can  -O"),
                ("905444444444", "13.07.2026", "10:02:00", "x", "00:00:01", "00:00:01", "can  -O"),
            ]
        )
        r2 = analyze_workbook(buf, min_repeat=2)[0]
        r3 = analyze_workbook(buf, min_repeat=3)[0]
        r4 = analyze_workbook(buf, min_repeat=4)[0]
        self.assertEqual(len(r2.tekrarli_numaralar), 1)
        self.assertEqual(len(r3.tekrarli_numaralar), 1)
        self.assertEqual(len(r4.tekrarli_numaralar), 0)

    def test_date_in_output(self):
        buf = _make_xlsx(
            [
                ("905555555555", "13.07.2026 00:00:00", "10:00:00", "x", "00:00:01", "00:00:01", "eda  -O"),
                ("905555555555", "14.07.2026 00:00:00", "11:00:00", "x", "00:00:01", "00:00:01", "eda  -O"),
            ]
        )
        n = analyze_workbook(buf, min_repeat=2)[0].tekrarli_numaralar[0]
        self.assertEqual(n.arama_sayisi, 2)
        self.assertTrue(any("13.07.2026" in s for s in n.saatler))
        self.assertTrue(any("14.07.2026" in s for s in n.saatler))
        self.assertIn("10:00:00", n.saatler[0])
        self.assertIn("11:00:00", n.saatler[1])

    def test_report_workbook_structure(self):
        buf = _make_xlsx(
            [
                ("905666666666", "13.07.2026", "10:00:00", "x", "00:00:01", "00:00:01", "naz  -O"),
                ("905666666666", "13.07.2026", "10:01:00", "x", "00:00:01", "00:00:01", "naz  -O"),
            ]
        )
        reports = analyze_workbook(buf, min_repeat=2)
        out = build_report_workbook(reports, min_repeat=2)
        wb = load_workbook(out)
        self.assertIn("Bilgi", wb.sheetnames)
        self.assertIn("Ozet", wb.sheetnames)
        self.assertIn("TekrarliAramalar", wb.sheetnames)
        self.assertIn("Naz", wb.sheetnames)
        ozet = wb["Ozet"]
        # header + 1 personel
        self.assertEqual(ozet.cell(2, 1).value, "Naz")
        self.assertEqual(ozet.cell(2, 2).value, 2)
        detay = wb["TekrarliAramalar"]
        self.assertEqual(detay.cell(2, 2).value, "905666666666")
        self.assertEqual(detay.cell(2, 3).value, 2)
        wb.close()


class TestBotHelpers(unittest.TestCase):
    def test_safe_filename(self):
        self.assertEqual(safe_filename("../../etc/passwd.xlsx"), "passwd.xlsx")
        self.assertTrue(safe_filename("rapor.xlsx").endswith(".xlsx"))
        self.assertNotIn("..", safe_filename("../x.xlsx"))
        self.assertTrue(safe_filename("a/b\\c.xlsx").endswith("c.xlsx"))

    def test_group_only_helper(self):
        from types import SimpleNamespace
        from telegram.constants import ChatType

        private = SimpleNamespace(
            effective_chat=SimpleNamespace(type=ChatType.PRIVATE)
        )
        group = SimpleNamespace(
            effective_chat=SimpleNamespace(type=ChatType.GROUP)
        )
        super_g = SimpleNamespace(
            effective_chat=SimpleNamespace(type=ChatType.SUPERGROUP)
        )
        none_chat = SimpleNamespace(effective_chat=None)

        self.assertFalse(is_group_chat(private))
        self.assertTrue(is_group_chat(group))
        self.assertTrue(is_group_chat(super_g))
        self.assertFalse(is_group_chat(none_chat))


@unittest.skipUnless(SAMPLE.exists(), "örnek Excel yok")
class TestSampleExcel(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.reports = analyze_workbook(SAMPLE, min_repeat=2)
        cls.by_name = {r.personel: r for r in cls.reports}

    def test_no_k_personnel(self):
        self.assertNotIn("K", self.by_name)
        self.assertNotIn("-K", self.by_name)

    def test_expected_personnel_count(self):
        # Örnekte 20 isimli personel (+ volkantest), -K hariç
        self.assertGreaterEqual(len(self.reports), 19)
        self.assertIn("Seda", self.by_name)
        self.assertIn("Celal", self.by_name)

    def test_celal_top_repeat(self):
        celal = self.by_name["Celal"]
        self.assertGreater(len(celal.tekrarli_numaralar), 0)
        top = celal.tekrarli_numaralar[0]
        self.assertGreaterEqual(top.arama_sayisi, 2)
        # Bilinen örnek: 905438447017 x3
        phones = {n.telefon: n.arama_sayisi for n in celal.tekrarli_numaralar}
        self.assertIn("905438447017", phones)
        self.assertEqual(phones["905438447017"], 3)

    def test_volkantest_nine_calls(self):
        v = self.by_name.get("Volkantest")
        self.assertIsNotNone(v)
        self.assertEqual(v.tekrarli_numaralar[0].arama_sayisi, 9)

    def test_summary_and_report_bytes(self):
        summary = format_text_summary(self.reports)
        self.assertIn("Tekrarlı Arama Özeti", summary)
        self.assertLessEqual(len(summary), 3600)
        data = build_report_workbook(self.reports).getvalue()
        self.assertGreater(len(data), 1000)
        # Geçerli xlsx
        wb = load_workbook(BytesIO(data))
        self.assertTrue(wb.sheetnames)
        wb.close()

    def test_write_temp_report(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "rapor.xlsx"
            path.write_bytes(build_report_workbook(self.reports).getvalue())
            self.assertTrue(path.exists())
            self.assertGreater(path.stat().st_size, 1000)


if __name__ == "__main__":
    unittest.main(verbosity=2)
