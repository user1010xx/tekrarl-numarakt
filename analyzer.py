"""
Excel arama kaydı analizi.

Personel bazlı: aynı numarayı farklı personeller arayabilir; tekrarlar
yalnızca aynı personel içinde sayılır.
"""

from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import Any, BinaryIO, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Sütun indeksleri (0-based): A, B, C, E, F, G
COL_TELEFON = 0
COL_TARIH = 1
COL_ARAMA_SAATI = 2
COL_KONUSMA = 4
COL_CALDIRMA = 5
COL_PERSONEL = 6

# Sadece "K", "-", "-K" gibi personel adı olmayanlar elenir.
# "seda  -O", "Ahmet-O" gibi isim içerenler alınır.
_INVALID_ONLY = re.compile(
    r"^[\s\-–—_/\\|]*[KkOo]?[\s\-–—_/\\|]*$"
)
_NAME_SUFFIX = re.compile(r"\s*[-–—]\s*[OoKk]\s*$")
_LETTER = re.compile(r"[A-Za-zÀ-ÿĞğÜüŞşİıÖöÇç]")
_TIME_HM = re.compile(r"^(\d{1,2}):(\d{2})(?::(\d{2}))?")


@dataclass
class CallRecord:
    telefon: str
    saat: str
    konusma_suresi: str
    caldirma_suresi: str
    konusma_saniye: int = 0
    caldirma_saniye: int = 0
    sort_key: str = ""


@dataclass
class NumberSummary:
    telefon: str
    arama_sayisi: int
    saatler: list[str] = field(default_factory=list)
    toplam_konusma_saniye: int = 0
    toplam_caldirma_saniye: int = 0
    konusma_detay: list[str] = field(default_factory=list)
    caldirma_detay: list[str] = field(default_factory=list)


@dataclass
class PersonnelReport:
    personel: str
    tekrarli_numaralar: list[NumberSummary] = field(default_factory=list)
    toplam_arama: int = 0
    tekrarli_arama_toplami: int = 0


def parse_duration_to_seconds(value: Any) -> int:
    """'00:00:56', timedelta, datetime veya saniye sayılarını saniyeye çevirir."""
    if value is None or value == "":
        return 0
    if isinstance(value, timedelta):
        return int(value.total_seconds())
    if isinstance(value, datetime):
        # Excel bazen süreyi datetime olarak okur
        return value.hour * 3600 + value.minute * 60 + value.second
    if isinstance(value, (int, float)):
        # Excel süre fraksiyonu (1 gün = 1.0) veya ham saniye
        if 0 < float(value) < 1:
            return int(round(float(value) * 86400))
        return int(value)

    text = str(value).strip()
    if not text or text.lower() in {"false", "none", "nan"}:
        return 0

    parts = text.split(":")
    try:
        if len(parts) == 3:
            h, m, s = int(parts[0]), int(parts[1]), int(float(parts[2]))
            return h * 3600 + m * 60 + s
        if len(parts) == 2:
            m, s = int(parts[0]), int(float(parts[1]))
            return m * 60 + s
        return int(float(text))
    except (ValueError, TypeError):
        return 0


def format_seconds(total: int) -> str:
    total = max(0, int(total))
    h, rem = divmod(total, 3600)
    m, s = divmod(rem, 60)
    return f"{h:02d}:{m:02d}:{s:02d}"


def normalize_phone(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return ""
    if isinstance(value, float):
        if value.is_integer():
            value = int(value)
        else:
            # Bilimsel / kesirli float numaraları reddet (yanlış birleştirme riski)
            return ""
    if isinstance(value, int):
        digits = str(abs(value))
        return digits if len(digits) >= 7 else ""

    text = str(value).strip()
    if not text or text.lower() in {"false", "none", "nan", "null", "true"}:
        return ""
    # Bilimsel gösterim string gelirse güvenli değil
    if re.search(r"[eE][+-]?\d+$", text):
        try:
            as_float = float(text)
            if as_float.is_integer():
                digits = str(int(as_float))
                return digits if len(digits) >= 7 else ""
        except ValueError:
            return ""
        return ""
    if text.endswith(".0") and text[:-2].replace("-", "").isdigit():
        text = text[:-2]
    digits = re.sub(r"\D", "", text)
    # En az 7 haneli numara kabul et (kısa/bozuk değerleri ele)
    if len(digits) < 7:
        return ""
    return digits


def format_time(value: Any) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return value.strftime("%H:%M:%S")
    if isinstance(value, timedelta):
        total = int(value.total_seconds()) % 86400
        return format_seconds(total)
    text = str(value).strip()
    # "13:15:57" veya "13:15:57.000"
    m = re.match(r"^(\d{1,2}:\d{2}:\d{2})", text)
    if m:
        return m.group(1)
    m2 = re.match(r"^(\d{1,2}:\d{2})$", text)
    if m2:
        return m2.group(1) + ":00"
    return text


def format_date(value: Any) -> str:
    """Tarihi kısa gösterime çevirir (gg.aa.yyyy)."""
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    text = str(value).strip()
    if not text or text.lower() in {"false", "none", "nan"}:
        return ""
    # "13.07.2026 00:00:00" -> "13.07.2026"
    m = re.match(r"^(\d{1,2}[./-]\d{1,2}[./-]\d{2,4})", text)
    if m:
        return m.group(1).replace("-", ".").replace("/", ".")
    # ISO
    m2 = re.match(r"^(\d{4})-(\d{2})-(\d{2})", text)
    if m2:
        return f"{m2.group(3)}.{m2.group(2)}.{m2.group(1)}"
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    return text[:10]


def combine_datetime_label(tarih: str, saat: str) -> str:
    if tarih and saat:
        return f"{tarih} {saat}"
    return saat or tarih or ""


def sort_key_for_call(tarih: str, saat: str) -> str:
    """Kronolojik sıralama anahtarı (YYYYMMDDHHMMSS)."""
    y, mo, d = "0000", "00", "00"
    if tarih:
        m = re.match(r"^(\d{1,2})\.(\d{1,2})\.(\d{2,4})$", tarih)
        if m:
            d = m.group(1).zfill(2)
            mo = m.group(2).zfill(2)
            y = m.group(3)
            if len(y) == 2:
                y = "20" + y
    hh, mm, ss = "00", "00", "00"
    if saat:
        tm = _TIME_HM.match(saat)
        if tm:
            hh = tm.group(1).zfill(2)
            mm = tm.group(2).zfill(2)
            ss = (tm.group(3) or "00").zfill(2)
    return f"{y}{mo}{d}{hh}{mm}{ss}"


def tr_lower(text: str) -> str:
    """Türkçe + İngilizce küçük harf (I→ı, İ→i)."""
    return (
        text.replace("İ", "i")
        .replace("I", "ı")
        .replace("Ş", "ş")
        .replace("Ğ", "ğ")
        .replace("Ü", "ü")
        .replace("Ö", "ö")
        .replace("Ç", "ç")
        .lower()
    )


def tr_upper(text: str) -> str:
    """Türkçe + İngilizce büyük harf (i→İ, ı→I)."""
    return (
        text.replace("i", "İ")
        .replace("ı", "I")
        .replace("ş", "Ş")
        .replace("ğ", "Ğ")
        .replace("ü", "Ü")
        .replace("ö", "Ö")
        .replace("ç", "Ç")
        .upper()
    )


def tr_capitalize_word(word: str) -> str:
    """Tek kelimeyi Türkçe kurallarla Başharf büyük yapar."""
    if not word:
        return word
    first = tr_upper(word[0])
    rest = tr_lower(word[1:]) if len(word) > 1 else ""
    return first + rest


def is_valid_personnel(raw: Any) -> bool:
    """Personel adı içeren kayıtları kabul eder; K, -, -K vb. reddeder."""
    if raw is None:
        return False
    text = str(raw).strip()
    if not text:
        return False
    if _INVALID_ONLY.match(text):
        return False
    base = _NAME_SUFFIX.sub("", text).strip()
    base = re.sub(r"^[\s\-–—]+|[\s\-–—]+$", "", base).strip()
    if not base:
        return False
    # K / O tek harf (büyük-küçük / TR duyarsız)
    if tr_upper(base) in {"K", "O"}:
        return False
    return bool(_LETTER.search(base))


def normalize_personnel_name(raw: Any) -> str:
    """
    Personel adını birleştirme için normalize eder.
    Türkçe (İ/ı/Ş/ğ…) ve İngilizce harfleri doğru büyük/küçük yapar.
    Örnek: 'seda  -O', 'SEDA -o', 'Seda-O' → 'Seda'
           'İREM -O', 'irem -O' → 'İrem'
    """
    text = str(raw).strip()
    base = _NAME_SUFFIX.sub("", text).strip()
    base = re.sub(r"\s+", " ", base).strip()
    if not base:
        return text
    parts = [p for p in base.split(" ") if p]
    return " ".join(tr_capitalize_word(p) for p in parts)


def _iter_data_rows(ws) -> Iterable[tuple]:
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        yield row


def _duration_display(raw: Any, seconds: int) -> str:
    if raw in (None, ""):
        return format_seconds(seconds)
    if isinstance(raw, (datetime, timedelta)):
        return format_seconds(seconds)
    text = str(raw).strip()
    if ":" not in text:
        return format_seconds(seconds)
    # "00:00:56.0" gibi
    m = re.match(r"^(\d{1,2}:\d{2}:\d{2})", text)
    if m:
        return m.group(1)
    return format_seconds(seconds)


def analyze_workbook(
    source: str | Path | BinaryIO | BytesIO,
    *,
    min_repeat: int = 2,
) -> list[PersonnelReport]:
    """
    Excel dosyasını personel bazlı analiz eder.

    min_repeat: Raporlanacak minimum arama sayısı (varsayılan 2 = tekrar).
    """
    if min_repeat < 1:
        min_repeat = 1

    wb = load_workbook(source, read_only=True, data_only=True)
    try:
        ws = wb.active
        # personel -> telefon -> [CallRecord]
        buckets: dict[str, dict[str, list[CallRecord]]] = defaultdict(
            lambda: defaultdict(list)
        )
        total_calls: dict[str, int] = defaultdict(int)

        for row in _iter_data_rows(ws):
            if len(row) <= COL_PERSONEL:
                continue
            personel_raw = row[COL_PERSONEL]
            if not is_valid_personnel(personel_raw):
                continue

            personel = normalize_personnel_name(personel_raw)
            telefon = normalize_phone(row[COL_TELEFON] if len(row) > COL_TELEFON else None)
            if not telefon:
                continue

            tarih = format_date(row[COL_TARIH] if len(row) > COL_TARIH else None)
            saat = format_time(row[COL_ARAMA_SAATI] if len(row) > COL_ARAMA_SAATI else None)
            label = combine_datetime_label(tarih, saat)
            konusma_raw = row[COL_KONUSMA] if len(row) > COL_KONUSMA else None
            caldirma_raw = row[COL_CALDIRMA] if len(row) > COL_CALDIRMA else None
            konusma_sn = parse_duration_to_seconds(konusma_raw)
            caldirma_sn = parse_duration_to_seconds(caldirma_raw)
            konusma_str = _duration_display(konusma_raw, konusma_sn)
            caldirma_str = _duration_display(caldirma_raw, caldirma_sn)

            buckets[personel][telefon].append(
                CallRecord(
                    telefon=telefon,
                    saat=label,
                    konusma_suresi=konusma_str,
                    caldirma_suresi=caldirma_str,
                    konusma_saniye=konusma_sn,
                    caldirma_saniye=caldirma_sn,
                    sort_key=sort_key_for_call(tarih, saat),
                )
            )
            total_calls[personel] += 1
    finally:
        wb.close()

    reports: list[PersonnelReport] = []
    for personel in sorted(buckets.keys(), key=lambda x: x.casefold()):
        numbers: list[NumberSummary] = []
        tekrar_toplam = 0
        for telefon, calls in buckets[personel].items():
            count = len(calls)
            if count < min_repeat:
                continue
            ordered = sorted(calls, key=lambda c: (c.sort_key, c.saat or ""))
            numbers.append(
                NumberSummary(
                    telefon=telefon,
                    arama_sayisi=count,
                    saatler=[c.saat for c in ordered if c.saat],
                    toplam_konusma_saniye=sum(c.konusma_saniye for c in ordered),
                    toplam_caldirma_saniye=sum(c.caldirma_saniye for c in ordered),
                    konusma_detay=[c.konusma_suresi for c in ordered],
                    caldirma_detay=[c.caldirma_suresi for c in ordered],
                )
            )
            tekrar_toplam += count

        numbers.sort(key=lambda n: (-n.arama_sayisi, n.telefon))
        reports.append(
            PersonnelReport(
                personel=personel,
                tekrarli_numaralar=numbers,
                toplam_arama=total_calls[personel],
                tekrarli_arama_toplami=tekrar_toplam,
            )
        )

    return reports


def build_report_workbook(reports: list[PersonnelReport], min_repeat: int = 2) -> BytesIO:
    """Personel bazlı Excel raporu üretir."""
    wb = Workbook()

    # --- Özet sayfa ---
    ws_ozet = wb.active
    ws_ozet.title = "Ozet"
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    alt_fill = PatternFill("solid", fgColor="D6EAF8")

    ozet_headers = [
        "Personel",
        "Toplam Arama",
        "Tekrarlı Numara Adedi",
        "Tekrarlı Arama Toplamı",
    ]
    ws_ozet.append(ozet_headers)
    for col, _ in enumerate(ozet_headers, 1):
        cell = ws_ozet.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font

    for rep in reports:
        ws_ozet.append(
            [
                rep.personel,
                rep.toplam_arama,
                len(rep.tekrarli_numaralar),
                rep.tekrarli_arama_toplami,
            ]
        )

    for col in range(1, 5):
        ws_ozet.column_dimensions[get_column_letter(col)].width = 24

    # --- Detay sayfa ---
    ws_detay = wb.create_sheet("TekrarliAramalar")
    detay_headers = [
        "Personel",
        "Telefon",
        "Arama Sayısı",
        "Tarih / Saatler",
        "Toplam Konuşma",
        "Toplam Çaldırma",
        "Konuşma Detay",
        "Çaldırma Detay",
    ]
    ws_detay.append(detay_headers)
    for col, _ in enumerate(detay_headers, 1):
        cell = ws_detay.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font

    row_idx = 2
    for rep in reports:
        if not rep.tekrarli_numaralar:
            continue
        for num in rep.tekrarli_numaralar:
            ws_detay.append(
                [
                    rep.personel,
                    num.telefon,
                    num.arama_sayisi,
                    ", ".join(num.saatler),
                    format_seconds(num.toplam_konusma_saniye),
                    format_seconds(num.toplam_caldirma_saniye),
                    ", ".join(num.konusma_detay),
                    ", ".join(num.caldirma_detay),
                ]
            )
            if row_idx % 2 == 0:
                for col in range(1, 9):
                    ws_detay.cell(row_idx, col).fill = alt_fill
            row_idx += 1

    widths = [16, 16, 14, 48, 16, 16, 30, 30]
    for i, w in enumerate(widths, 1):
        ws_detay.column_dimensions[get_column_letter(i)].width = w
        for cell in ws_detay[get_column_letter(i)]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # --- Personel sayfaları (kısa ad) ---
    used_titles: set[str] = set()
    for rep in reports:
        if not rep.tekrarli_numaralar:
            continue
        title = _safe_sheet_title(rep.personel, used_titles)
        ws = wb.create_sheet(title)
        ws.append(
            [
                "Telefon",
                "Arama Sayısı",
                "Tarih / Saatler",
                "Toplam Konuşma",
                "Toplam Çaldırma",
                "Konuşma Detay",
                "Çaldırma Detay",
            ]
        )
        for col in range(1, 8):
            cell = ws.cell(1, col)
            cell.fill = header_fill
            cell.font = header_font
        for num in rep.tekrarli_numaralar:
            ws.append(
                [
                    num.telefon,
                    num.arama_sayisi,
                    ", ".join(num.saatler),
                    format_seconds(num.toplam_konusma_saniye),
                    format_seconds(num.toplam_caldirma_saniye),
                    ", ".join(num.konusma_detay),
                    ", ".join(num.caldirma_detay),
                ]
            )
        for i, w in enumerate([16, 14, 48, 16, 16, 30, 30], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # Bilgi sayfası
    ws_info = wb.create_sheet("Bilgi", 0)
    ws_info["A1"] = "Tekrarlı Arama Raporu"
    ws_info["A1"].font = Font(bold=True, size=14)
    ws_info["A3"] = "Kurallar"
    ws_info["A3"].font = Font(bold=True)
    ws_info["A4"] = (
        f"• Aynı personelin aynı numarayı en az {min_repeat} kez araması tekrar sayılır."
    )
    ws_info["A5"] = (
        "• Farklı personellerin aynı numarayı araması birbirinden bağımsızdır."
    )
    ws_info["A6"] = (
        "• G sütununda personel adı olmayan kayıtlar (K, -, -K vb.) hariç tutulur."
    )
    ws_info["A7"] = (
        "• Sütunlar: A=Telefon, B=Tarih, C=Arama Saati, E=Konuşma, F=Çaldırma, G=Personel"
    )
    ws_info.column_dimensions["A"].width = 90

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _safe_sheet_title(name: str, used: set[str]) -> str:
    # Excel sheet adı max 31 karakter, yasak: \ / * ? : [ ]
    cleaned = re.sub(r'[\\/*?:\[\]]', "", name).strip() or "Personel"
    cleaned = cleaned[:28]
    base = cleaned
    i = 1
    while cleaned in used:
        cleaned = f"{base[:25]}_{i}"
        i += 1
    used.add(cleaned)
    return cleaned


def format_text_summary(reports: list[PersonnelReport], *, max_chars: int = 3500) -> str:
    """Telegram mesajı için düz metin özet (Markdown yok — parse hatası riski yok)."""
    lines: list[str] = ["Tekrarlı Arama Özeti", ""]
    aktif = [r for r in reports if r.tekrarli_numaralar]
    lines.append(f"Personel sayısı: {len(reports)}")
    lines.append(f"Tekrarlı araması olan: {len(aktif)}")
    lines.append("")

    for rep in aktif:
        lines.append(f"• {rep.personel}")
        lines.append(
            f"  Toplam arama: {rep.toplam_arama} | "
            f"Tekrarlı numara: {len(rep.tekrarli_numaralar)}"
        )
        for num in rep.tekrarli_numaralar[:5]:
            saat_txt = ", ".join(num.saatler[:6])
            if len(num.saatler) > 6:
                saat_txt += f" …(+{len(num.saatler) - 6})"
            lines.append(
                f"  - {num.telefon} → {num.arama_sayisi}x\n"
                f"    Saat: {saat_txt}\n"
                f"    Konuşma: {format_seconds(num.toplam_konusma_saniye)} | "
                f"Çaldırma: {format_seconds(num.toplam_caldirma_saniye)}"
            )
        if len(rep.tekrarli_numaralar) > 5:
            lines.append(
                f"  … ve {len(rep.tekrarli_numaralar) - 5} numara daha (Excel'de)"
            )
        lines.append("")

    if not aktif:
        lines.append("Belirtilen eşiğin üzerinde tekrarlı arama bulunamadı.")

    text = "\n".join(lines).strip()
    if len(text) > max_chars:
        text = text[: max_chars - 40] + "\n\n… (devamı Excel dosyasında)"
    return text
